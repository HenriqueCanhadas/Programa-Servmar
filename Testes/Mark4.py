import tkinter as tk
from tkinter import filedialog
import xlwings as xw
import openpyxl
import pandas as pd
import customtkinter 
import threading
from PIL import Image
from tkinter import font
import os

file_path = ""

#-----------------------------

def carregar_programa():
    janela.iconbitmap(icone_path)
    thread = threading.Thread(target=carregar_excel)
    thread.start()

def iniciar_programa():
    janela.iconbitmap(icone_path)
    thread = threading.Thread(target=rodar_programa)
    thread.start()

def salvar_programa():
    janela.iconbitmap(icone_path)
    thread = threading.Thread(target=salvar_excel)
    thread.start()

#-----------------------------

def carregar_excel():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if file_path:
        global dataframe
        dataframe = pd.read_excel(file_path)
        status_label.config(text="Arquivo Excel Carregado.", foreground="black")
        print("Arquivo Excel Carregado.")
        print("-" * 100)

def rodar_programa():
    global file_path 
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl
    global valores_cliente

    if file_path == "":
        status_label.config(text="Nenhum Excel Carregado!", foreground="red")
        print("Nenhum Excel Carregado!")
        print("-" * 100)
        return
    
    status_label.config(text="Rodando o Programa", foreground="black")
    janela.update() 

    wb_xlwings = xw.Book(file_path) 

    valores_cliente=[]
    valores_xlwings = []
    valores_workbook_openpyxl = []

    planilha = wb_xlwings.sheets['DADOS DO CLIENTE']

    planilha.activate()

    razao = planilha.range('B6').value
    endereco = planilha.range('B8').value
    contato= planilha.range('B10').value
    telefone = planilha.range('G10').value
    projeto = planilha.range('C14').value
    numprojeto = planilha.range('J14').value
    endeprojeto = planilha.range('C16').value
    numos= planilha.range('J18').value
    

    valores_cliente.append((razao,endereco,contato,telefone,projeto,numprojeto,endeprojeto,numos))

    print(f'Valores lidos em {planilha}:')
    print(f'Razão: {razao}')
    print(f'Endereço: {endereco}')
    print(f'Contato: {contato}')
    print(f'Telefone: {telefone}')
    print(f'Nome do Projeto: {projeto}')
    print(f'N° Projeto: {numprojeto}')
    print(f'Endereço Projeto: {endeprojeto}')
    print(f'N° OS: {numos}')
    print("-" * 100)

    for sheet_name in wb_xlwings.sheets:
        planilha = wb_xlwings.sheets[sheet_name]

        planilha.activate()

        condutividade = planilha.range('E41').value
        oxirreducao = planilha.range('G41').value
        oxigenio = planilha.range('I41').value
        ph = planilha.range('L41').value
        temperatura = planilha.range('N41').value
        turbidez = planilha.range('P41').value
        desvio1 = planilha.range('A65').value 
        desvio2 = planilha.range('A66').value 
        desvio3 = planilha.range('A67').value
        
        valores_xlwings.append((condutividade,oxirreducao,oxigenio, ph,temperatura,turbidez,desvio1,desvio2,desvio3))

        print(f'Valores lidos em {sheet_name}:')
        print(f'Condutividade: {condutividade}')
        print(f'Potencial de oxirredução: {oxirreducao}')
        print(f'Oxigênio Dissolvido: {oxigenio}')
        print(f'pH: {ph}')
        print(f'Temperatura: {temperatura}')
        print(f'Turbidez: {turbidez}')
        print(f'Desvio1: {desvio1}')
        print(f'Desvio2: {desvio2}')
        print(f'Desvio3: {desvio3}')

        print("-" * 100)

    wb_xlwings.close()

    workbook_openpyxl = openpyxl.load_workbook(file_path)

    for sheet_name in workbook_openpyxl.sheetnames:
        planilha_openpyxl = workbook_openpyxl[sheet_name]

        identificacao = planilha_openpyxl['C30'].value
        data = planilha_openpyxl['C6'].value
        horaensaio = planilha_openpyxl['H30'].value
        horaamostragem = planilha_openpyxl['K30'].value
        condicoes = planilha_openpyxl['P30'].value

        observacao = planilha_openpyxl['A69'].value
        multiparametro = planilha_openpyxl['I14'].value
        bomba = planilha_openpyxl['I16'].value
        turbidimetro = planilha_openpyxl['I18'].value
        medidor = planilha_openpyxl['I20'].value
        painel = planilha_openpyxl['I22'].value
        termometro = planilha_openpyxl['S18'].value
        vazao= planilha_openpyxl['F30'].value
        realizador = planilha_openpyxl['D8'].value

        valores_workbook_openpyxl.append((identificacao,data, horaensaio,horaamostragem,condicoes,observacao,multiparametro,bomba,turbidimetro,medidor,painel,termometro,vazao,realizador))
        
        print(f'Identificação/Aba/Guia: {identificacao}')
        print(f'Data: {data}')
        print(f'Hora do ensaio: {horaensaio}')
        print(f'Hora da amostragem: {horaamostragem}')
        print(f'Condições ambientais: {condicoes}')
        print(f'Observações: {observacao}')
        print(f'Multiparâmetro: {multiparametro}')
        print(f'Bomba: {bomba}')
        print(f'Turbidimetro: {turbidimetro}')
        print(f'Medidor de Nível: {medidor}')
        print(f'Painel controlador: {painel}')
        print(f'Termometro: {termometro}')
        print(f'Vazão: {vazao}')
        print(f'Amostragem e ensaio realizado por: {realizador}')

        print("-" * 100) 

    status_label.config(text="Todos os dados foram Armazenados", foreground="black")

    print('Todos os dados foram Armazenados')

    print("-" * 100)

def salvar_excel():
    global valores_cliente
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl
    global file_path
    global status_label

    if file_path == "":
        status_label.config(text="Nenhum Excel Carregado!", foreground="red")
        print("Nenhum Excel Carregado!")
        print("-" * 100)
        return
    
    try:
        if valores_cliente is None:
            status_label.config(text="Programa não foi Rodado!", foreground="red")
            print("Programa não foi Rodado!")
            print("-" * 100)
            print("teste")
            return
    except NameError:
        status_label.config(text="Programa não foi Rodado!", foreground="red")
        print("Programa não foi Rodado!")
        print("-" * 100)
        return

    linha=72
    while True:
        dialog = customtkinter.CTkInputDialog(title="Caixa de dialogo", text="Digite a partir de qual linha seja inserirdo os dados:")
        linha_text = dialog.get_input()

        try:        
            if linha_text is None:
                break
            linha = int(linha_text)
            break  # Sai do loop se a conversão for bem-sucedida
        except ValueError as e:
            print("O valor não pode ser convertido para um número inteiro. Tente novamente.")
    
    if file_path:
        status_label.config(text="Preparando para salvar o arquivo.")

        print(f"Abrindo o arquivo Excel existente na pasta:{file_path}")
        workbook_openpyxl = openpyxl.load_workbook(file_path)

        sheet = workbook_openpyxl['FINAL']

        while len(valores_cliente) > 0:
            if len(valores_cliente) > 0:
                razao,endereco,contato, telefone, projeto,numprojeto,endeprojeto,numos=valores_cliente.pop(0)
                
                sheet['B12'] = razao
                sheet['B14'] = endereco
                sheet['B16'] = contato
                sheet['G16'] = telefone
                sheet['C20'] = projeto
                sheet['J20'] = numprojeto
                sheet['C22'] = endeprojeto
                sheet['J24'] = numos

        while len(valores_xlwings) > 0 and len(valores_workbook_openpyxl) > 0:
            if len(valores_xlwings) > 0:
                condutividade, oxirreducao, oxigenio, ph, temperatura, turbidez, desvio1,desvio2,desvio3 = valores_xlwings.pop(0)

                sheet[f'E{linha}'] = condutividade
                sheet[f'F{linha}'] = oxirreducao
                sheet[f'G{linha}'] = oxigenio
                sheet[f'H{linha}'] = ph
                sheet[f'I{linha}'] = temperatura
                sheet[f'J{linha}'] = turbidez

                sheet[f'L{linha}'] = desvio1
                sheet[f'M{linha}'] = desvio2
                sheet[f'N{linha}'] = desvio3
  
            if len(valores_workbook_openpyxl) > 0:
                identificacao, data, horaensaio, horaamostragem, condicoes, observacao, multiparametro, bomba, turbidimetro, medidor, painel, termometro, vazao, realizador = valores_workbook_openpyxl.pop(0)

                sheet[f'A{linha}'] = identificacao
                sheet[f'B{linha}'] = data
                sheet[f'C{linha}'] = horaensaio
                sheet[f'D{linha}'] = horaamostragem
                sheet[f'K{linha}'] = condicoes

                sheet[f'O{linha}'] = observacao
                sheet[f'P{linha}'] = multiparametro
                sheet[f'Q{linha}'] = bomba
                sheet[f'R{linha}'] = turbidimetro
                sheet[f'S{linha}'] = medidor
                sheet[f'T{linha}'] = painel
                sheet[f'U{linha}'] = termometro
                sheet[f'V{linha}'] = vazao
                sheet[f'W{linha}'] = realizador
            
            linha += 1

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if save_path:
        workbook_openpyxl.save(save_path)
        status_label.config(text="Excel salvo com sucesso!", foreground="black")

        print("Excel salvo com sucesso!")
        print("-" * 100)

#-----------------------------

def imagem():
  global janela

  imagem_path = os.path.join(script_dir, "servmarlogo.png")
  img = customtkinter.CTkImage(light_image=Image.open(imagem_path), size=(345, 50))
  customtkinter.CTkButton(janela,text="",state="disable",fg_color="transparent",image=img).pack(pady=5)
    
def main():
    global status_label
    global janela
    global icone_path
    global script_dir

    # Obtém o diretório atual do script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    icone_path = os.path.join(script_dir, "servmarico.ico")
    #Caso For rodar separado alterar para janela = customtkinter.CTk()
    janela = customtkinter.CTkToplevel()
    janela.geometry("335x350")
    janela.title("SERVMAR")
    janela.resizable(width=False, height=False)
    
    customtkinter.set_appearance_mode("light")
    customtkinter.set_default_color_theme("dark-blue")

    cabecalho= customtkinter.CTkLabel(janela, text="RELATÓRIO DE ENSAIO E AMOSTRAGEM",font=("arial bold", 16),text_color="#0000ff")
    cabecalho.pack()

    load_button = customtkinter.CTkButton(janela, text="Carregar Excel", command=carregar_programa)
    load_button.pack(pady=20)

    run_button = customtkinter.CTkButton(janela, text="Rodar Programa", command=iniciar_programa)
    run_button.pack(pady=20)

    save_button = customtkinter.CTkButton(janela, text="Salvar Excel", command=salvar_programa)
    save_button.pack(pady=20)

    status_label = tk.Label(janela, text="",bg="#f2f2f2", font=("Arial", 15))
    status_label.pack(pady=5)
    fonte = font.nametofont("TkDefaultFont")
    fonte.configure(underline=True)
    fonte.configure(size=15)
    status_label.config(font=fonte)

    imagem()

if __name__ == "__main__":
    main()